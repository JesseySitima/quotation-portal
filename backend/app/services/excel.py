from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font


def create_quotation_excel(quotation, items):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Quotation Items"

    # Headers
    headers = ["Product", "Quantity", "Unit"]

    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(
            row=1,
            column=column,
            value=header,
        )
        cell.font = Font(bold=True)

    # Items
    row = 2

    for item in items:
        worksheet.cell(
            row=row,
            column=1,
            value=item["product_name"],
        )

        worksheet.cell(
            row=row,
            column=2,
            value=item["quantity"],
        )

        worksheet.cell(
            row=row,
            column=3,
            value=item["unit"],
        )

        row += 1

    # Column widths
    worksheet.column_dimensions["A"].width = 40
    worksheet.column_dimensions["B"].width = 15
    worksheet.column_dimensions["C"].width = 20

    # Save to memory
    output = BytesIO()
    workbook.save(output)

    output.seek(0)

    return output.getvalue()